import asyncio
import base64
from concurrent.futures import ThreadPoolExecutor
import io
import json
import math
import os
import tempfile
import time
import uuid
import threading
import pandas as pd

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo
import boto3
from botocore.exceptions import ClientError
from dotenv import load_dotenv
import numpy as np
from openai import OpenAI
import requests
import streamlit as st
from PIL import Image

try:
    from google import genai  # Preferred client API
    _HAS_GOOGLE_CLIENT = True
except ImportError:  # Fall back to legacy generative AI module
    import google.generativeai as genai
    _HAS_GOOGLE_CLIENT = False
from langchain_core.prompts import PromptTemplate
from openpyxl import Workbook, load_workbook


load_dotenv()


def _get_env_bool(key: str, default: bool) -> bool:
    raw = os.environ.get(key)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "on"}


def _get_env_int(key: str, default: int) -> int:
    raw = os.environ.get(key)
    if raw is None:
        return default
    try:
        return int(raw)
    except ValueError:
        print(f"[warn] Invalid value for {key} ({raw!r}); using default {default}.")
        return default


MAX_PROMPT_CHARS = _get_env_int("MAX_PROMPT_CHARS", 32000)
S3_ALLOW_OBJECT_ACL = _get_env_bool("S3_ALLOW_OBJECT_ACL", True)
_S3_ACL_SUPPORTED = S3_ALLOW_OBJECT_ACL
_S3_ACL_WARNED = False
EXCEL_LOCK = threading.Lock()
GENERATION_LOCK = threading.Lock()


async def run_blocking(func, *args, **kwargs):
    return await asyncio.to_thread(func, *args, **kwargs)

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise ValueError("OPENAI_API_KEY must be set.")
OPENAI_CLIENT = OpenAI(api_key=OPENAI_API_KEY)

GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY")
if not GOOGLE_API_KEY:
    raise ValueError("GOOGLE_API_KEY must be set.")

if _HAS_GOOGLE_CLIENT:
    gclient = genai.Client(api_key=GOOGLE_API_KEY)
else:
    genai.configure(api_key=GOOGLE_API_KEY)

    class _GenAIFallbackClient:
        """Provide minimal Client(models.generate_content) surface."""

        class _ModelsProxy:
            def generate_content(self, model: str, contents: List[Any]):
                generative_model = genai.GenerativeModel(model)
                return generative_model.generate_content(contents)

        def __init__(self) -> None:
            self.models = self._ModelsProxy()

    gclient = _GenAIFallbackClient()

REPLICATE_API_TOKEN = os.environ.get("REPLICATE_API_TOKEN") or os.environ.get("replicate")
REPLICATE_MODEL_IDS = {
    "qwen/qwen-image-edit": "qwen/qwen-image-edit",
    "qwen/qwen-image-edit-plus": "qwen/qwen-image-edit-plus",
    "black-forest-labs/flux-pro": "black-forest-labs/flux-pro",
    "black-forest-labs/flux-kontext-max": "black-forest-labs/flux-kontext-max",
    "black-forest-labs/flux-kontext-pro": "black-forest-labs/flux-kontext-pro",
    "ideogram-ai/ideogram-v3-turbo": "ideogram-ai/ideogram-v3-turbo",
    "qwen/qwen-image": "qwen/qwen-image",
}

ARK_API_KEY = os.environ.get("ARK_API_KEY") or None

SEEDREAM_ENDPOINT = "https://ark.ap-southeast.bytepluses.com/api/v3/images/generate"

try:
    from byteplussdkarkruntime import Ark as BytePlusArk
    from byteplussdkarkruntime.types.images.images import SequentialImageGenerationOptions
except Exception as exc:  # noqa: BLE001
    BytePlusArk = None
    SequentialImageGenerationOptions = None
    if ARK_API_KEY:
        print("[warn] BytePlus Ark SDK unavailable; falling back to HTTP calls.", exc)

ARK_SDK_CLIENT = None
if BytePlusArk and ARK_API_KEY:
    try:
        ARK_SDK_CLIENT = BytePlusArk(
            base_url="https://ark.ap-southeast.bytepluses.com/api/v3",
            api_key=ARK_API_KEY,
        )
        print("[info] Seedream SDK client initialized.")
    except Exception as exc:  # noqa: BLE001
        ARK_SDK_CLIENT = None
        print("[warn] Failed to initialize Seedream SDK client; using HTTP fallback:", exc)

AWS_REGION = os.environ.get("AWS_REGION") or boto3.session.Session().region_name or "us-east-1"
S3_BUCKET_NAME = os.environ.get("S3_BUCKET_NAME")
S3_BASE_PREFIX = os.environ.get("S3_BASE_PREFIX", "lifestyle-app")
S3_INPUT_PREFIX = f"{S3_BASE_PREFIX}/inputs"
S3_OUTPUT_PREFIX = f"{S3_BASE_PREFIX}/outputs"
LOG_RUNS_PREFIX = os.environ.get("S3_LOG_RUNS_PREFIX", f"{S3_BASE_PREFIX}/logs/runs")
LOG_FEEDBACK_PREFIX = os.environ.get("S3_LOG_FEEDBACK_PREFIX", f"{S3_BASE_PREFIX}/logs/feedback")
DEFAULT_INFOGRAPHICS_EXCEL = f"{S3_BASE_PREFIX}/logs/1_infographics_llm_eval.xlsx"
S3_EXCEL_KEY = (
    os.environ.get("S3_INFOGRAPHICS_EXCEL_KEY")
    or os.environ.get("S3_EXCEL_KEY")
    or DEFAULT_INFOGRAPHICS_EXCEL
)

if not S3_BUCKET_NAME:
    raise ValueError("S3_BUCKET_NAME must be set for storage.")

s3_client = boto3.client("s3", region_name=AWS_REGION)


def s3_object_url(key: str) -> str:
    if AWS_REGION == "us-east-1":
        return f"https://{S3_BUCKET_NAME}.s3.amazonaws.com/{key}"
    return f"https://{S3_BUCKET_NAME}.s3.{AWS_REGION}.amazonaws.com/{key}"


def _acl_not_supported(exc: Exception) -> bool:
    message = str(exc)
    if isinstance(exc, ClientError):
        code = exc.response.get("Error", {}).get("Code", "")
        if code in {"AccessControlListNotSupported", "InvalidRequest"}:
            return True
    return "AccessControlListNotSupported" in message or "BucketOwnerEnforced" in message


def _maybe_disable_acl(extra_args: Dict[str, Any], exc: Exception) -> bool:
    global _S3_ACL_SUPPORTED, _S3_ACL_WARNED
    if not extra_args.get("ACL") or not _S3_ACL_SUPPORTED:
        return False
    if not _acl_not_supported(exc):
        return False
    extra_args.pop("ACL", None)
    _S3_ACL_SUPPORTED = False
    if not _S3_ACL_WARNED:
        print(
            "[warn] S3 bucket does not allow object ACLs; retrying uploads without ACL. "
            "Ensure a bucket policy or presigned URLs grant read access if needed."
        )
        _S3_ACL_WARNED = True
    return True


def upload_file_to_s3(local_path: Path, key_prefix: str, content_type: str = "image/jpeg", public: bool = True) -> Optional[str]:
    key = f"{key_prefix.strip('/')}/{datetime.now(ZoneInfo('UTC')).strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}{Path(local_path).suffix}"
    extra_args: Dict[str, Any] = {"ContentType": content_type}
    if public and _S3_ACL_SUPPORTED:
        extra_args["ACL"] = "public-read"
    try:
        s3_client.upload_file(str(local_path), S3_BUCKET_NAME, key, ExtraArgs=extra_args)
        return s3_object_url(key)
    except Exception as exc:  # noqa: BLE001
        if _maybe_disable_acl(extra_args, exc):
            try:
                s3_client.upload_file(str(local_path), S3_BUCKET_NAME, key, ExtraArgs=extra_args)
                return s3_object_url(key)
            except Exception as retry_exc:  # noqa: BLE001
                print("[error] Failed to upload file to S3 even without ACL:", retry_exc)
                return None
        print("[error] Failed to upload file to S3:", exc)
        return None


def download_s3_file(key: str, destination: Path) -> bool:
    try:
        destination.parent.mkdir(parents=True, exist_ok=True)
        s3_client.download_file(S3_BUCKET_NAME, key, str(destination))
        return True
    except ClientError as exc:
        if exc.response.get("Error", {}).get("Code") == "404":
            return False
        raise


def upload_bytes_to_s3(data: bytes, key: str, content_type: str = "application/octet-stream", public: bool = True) -> Optional[str]:
    extra_args: Dict[str, Any] = {"ContentType": content_type}
    if public and _S3_ACL_SUPPORTED:
        extra_args["ACL"] = "public-read"
    try:
        s3_client.put_object(Bucket=S3_BUCKET_NAME, Key=key, Body=data, **extra_args)
        return s3_object_url(key)
    except Exception as exc:  # noqa: BLE001
        if _maybe_disable_acl(extra_args, exc):
            try:
                s3_client.put_object(Bucket=S3_BUCKET_NAME, Key=key, Body=data, **extra_args)
                return s3_object_url(key)
            except Exception as retry_exc:  # noqa: BLE001
                print("[error] Failed to upload bytes to S3 even without ACL:", retry_exc)
                return None
        print("[error] Failed to upload bytes to S3:", exc)
        return None


def clamp_prompt_length(prompt: str, max_chars: int = MAX_PROMPT_CHARS) -> str:
    return prompt


EXCEL_HEADERS = [
    "Timestamp",
    "Category",
    "Description",
    "Model Name",
    "Output Image No",
    "Thumbs (1=Like,0=Dislike)",
    "Review",
    "Score by Human",
    "Input Image View Link",
    "Input Image Download Link",
    "Output Image View Link",
    "Output Image Download Link",
    "Inference Speed (img/sec)",
    "Latency (sec)",
    "Token Processed (in+out)",
    "Images Generated",
    "Image Quality (resolution)",
]


def create_excel_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(EXCEL_HEADERS)
    wb.save(path)


def download_excel_from_s3() -> Path:
    tmp_path = Path(tempfile.gettempdir()) / f"lifestyle_log_{uuid.uuid4().hex}.xlsx"
    if download_s3_file(S3_EXCEL_KEY, tmp_path):
        return tmp_path
    create_excel_template(tmp_path)
    s3_client.upload_file(str(tmp_path), S3_BUCKET_NAME, S3_EXCEL_KEY)
    return tmp_path


def upload_excel_to_s3(path: Path) -> None:
    s3_client.upload_file(str(path), S3_BUCKET_NAME, S3_EXCEL_KEY)


def append_row_to_excel(row: List[Any]) -> None:
    with EXCEL_LOCK:
        tmp = download_excel_from_s3()
        try:
            wb = load_workbook(tmp)
            ws = wb.active
            ws.append(row)
            wb.save(tmp)
            upload_excel_to_s3(tmp)
        finally:
            tmp.unlink(missing_ok=True)

MAIN_PROMPT = """
You are an Elite AI Art Director, Ecommerce Conversion Strategist, and Photorealistic Infographic Designer specializing in high-conversion Amazon/Etsy/Shopee/D2C infographic visuals. Your job is to generate one single photorealistic, premium-quality infographic image using the provided product reference image + product description + category — with absolute visual accuracy and zero alteration of product details.

Your infographic must instantly build trust, clarity, and desire, while maintaining 100% product fidelity.

---------------------------------------------------
INTERNAL CREATIVE REASONING (DO NOT OUTPUT)
1. Define the core conversion objective for this infographic.
2. Identify the customer problem being solved.
3. Evaluate visible product materials, geometry, colors, branding, textures, scale.
4. Identify key differentiators, emotional triggers, and proof signals.
5. Architect the content hierarchy: headline → hero visual → benefits → proof → CTA.
6. Validate realism: lighting, perspective, materials, shadows, scale.
7. Ensure all text placement avoids covering critical product features.
8. Map the desired emotional response (trust, clarity, excitement, safety, premium quality).

---------------------------------------------------
STRICT PRODUCT FIDELITY RULES (NON-NEGOTIABLE)
- Do not change even 0.1% of the product.
- No modification of shape, geometry, logo, color tones, materials, textures, size, proportions, or surface details.
- No stylization, warping, stretching, or over-editing.
- Product must be fully visible, unobstructed, and photorealistic.

---------------------------------------------------
PRODUCT ANALYSIS (INTERNAL, DO NOT OUTPUT)
Extract:
- Form/shape, materials, textures, visible features, colors.
- Product category and purpose.
- Any measurable elements (capacity, wattage, dimensions).
- Contextual cues (environment, lighting direction, scale).

---------------------------------------------------
STRATEGY DEVELOPMENT (INTERNAL, DO NOT OUTPUT)
1. Identify buyer persona and emotional drivers.
2. Create one-sentence value proposition (functional + emotional).
3. Build message hierarchy.
4. Position the product as premium, innovative, durable, sustainable, stylish, or reliable.

---------------------------------------------------
INFOGRAPHIC CONTENT FRAMEWORK

HEADLINE (≤ 10 words)
- Benefit-driven, premium, instantly clear.

SUBHEADLINE (≤ 18 words)
- Clarifies the promise and emotional/functional impact.

BENEFIT BULLETS (3–6 max)
- Outcome-focused statements such as:
  "Engineered for all-day comfort"
  "Built for long-lasting durability"
  "Precise performance you can trust"

OPTIONAL PROOF POINTS
- 4.8★ rated, 50,000+ sold, lab-tested, certified safe.

CTA (3–6 words)
- “Shop Now”, “Get Yours Today”, “Upgrade Now”.

---------------------------------------------------
VISUAL DESIGN SYSTEM

COMPOSITION & LAYOUT
- Eye flow: Headline → Product Hero → Benefits → Proof → CTA.
- Clean negative space; balanced alignment; no clutter.

TYPOGRAPHY
- Use modern sans-serif (Inter / Helvetica / Roboto / SF Pro).
- Hierarchy: Headline (48–60px), Subheadline (22–26px), Body (16–18px).
- Strong contrast; crisp readability.

COLOR PALETTE
- Derived from the product + clean neutrals.
- One strategic accent color for clarity and CTAs.

GRAPHIC ELEMENTS
- Minimal icons, thin dividers, subtle arrows.
- Flat or semi-flat style only.

LIGHTING & EXECUTION
- Photorealistic lighting (35–50mm perspective).
- Correct shadows, reflections, realistic textures.
- No CGI look, no surreal flares, no artificial glow.

---------------------------------------------------
STORYTELLING ELEMENT
Visually communicate why the product is worth buying by emphasizing:
- Trust
- Clarity
- Emotional resonance
- Demonstrated usefulness
- Realistic environment when appropriate

The viewer should think:
"This looks real, premium, and worth buying."

---------------------------------------------------
REALISM & ACCURACY VALIDATION
Check:
- No changes to product design or color.
- Lighting direction consistent.
- Realistic shadows.
- Text legible at thumbnail view.
- No distortion or oversaturation.
- Feels like a modern premium ecommerce graphic.

---------------------------------------------------
FINAL OUTPUT REQUIREMENTS
- Generate ONE 4K or 8K infographic image.
- Orientation: 4:5 or 16:9.
- Photorealistic, premium, clean, credible.
- No variations, no filters, no CGI artifacts.
- No added props that alter product meaning.

---------------------------------------------------
CREATIVE GOAL
Produce a conversion-optimized, photorealistic infographic that blends:
- Realism
- Trust
- Clarity
- Emotional appeal
- Premium branding
- 100% accuracy

The viewer should instantly feel:
“This is real, premium, trustworthy — and exactly what I need.”
"""

PROMPT_TEMPLATE = PromptTemplate.from_template(
    "{main_prompt}\\n\\nCategory: {category}\\nDescription: {description}"
).partial(main_prompt=MAIN_PROMPT)

PROMPT_TEMPLATE_SEEDREAM = PromptTemplate.from_template(
    "{main_prompt}\n\nCategory: {category}\nDescription: {description}"
).partial(main_prompt=MAIN_PROMPT)

MODELS = [
    "gpt-image-1",
    "gpt-image-1-mini",
    "gemini-2.5-flash-image",
    "gemini-3-pro-image-preview",
    "qwen/qwen-image-edit",
    "qwen/qwen-image-edit-plus",
    "black-forest-labs/flux-kontext-max",
    "black-forest-labs/flux-kontext-pro",
    "black-forest-labs/flux-pro",
    "ideogram-ai/ideogram-v3-turbo",
    "qwen/qwen-image",
    "seedream-4",
    "seededit-3-0-i2i-250628",
]

# Optional per-model prompt overrides. Leave entries empty or remove keys to use the main prompt.
MODEL_PROMPT_OVERRIDES: Dict[str, str] = {}
# Uncomment and set any of the lines below to override that model's prompt:
# MODEL_PROMPT_OVERRIDES["gpt-image-1"] = "Custom prompt for gpt-image-1"
# MODEL_PROMPT_OVERRIDES["gpt-image-1-mini"] = "Custom prompt for gpt-image-1-mini"
# MODEL_PROMPT_OVERRIDES["gemini-2.5-flash-image"] = "Custom prompt for gemini-2.5-flash-image"
# MODEL_PROMPT_OVERRIDES["gemini-3-pro-image-preview"] = "Custom prompt for gemini-3-pro-image-preview"
# MODEL_PROMPT_OVERRIDES["qwen/qwen-image-edit"] = "Custom prompt for qwen/qwen-image-edit"
# MODEL_PROMPT_OVERRIDES["qwen/qwen-image-edit-plus"] = "Custom prompt for qwen/qwen-image-edit-plus"
# MODEL_PROMPT_OVERRIDES["black-forest-labs/flux-kontext-max"] = "Custom prompt for black-forest-labs/flux-kontext-max"
# MODEL_PROMPT_OVERRIDES["black-forest-labs/flux-kontext-pro"] = "Custom prompt for black-forest-labs/flux-kontext-pro"
# MODEL_PROMPT_OVERRIDES["black-forest-labs/flux-pro"] = "Custom prompt for black-forest-labs/flux-pro"
# MODEL_PROMPT_OVERRIDES["ideogram-ai/ideogram-v3-turbo"] = "Custom prompt for ideogram-ai/ideogram-v3-turbo"
# MODEL_PROMPT_OVERRIDES["qwen/qwen-image"] = "Custom prompt for qwen/qwen-image"
# MODEL_PROMPT_OVERRIDES["seedream-4"] = "Custom prompt for seedream-4"
# MODEL_PROMPT_OVERRIDES["seededit-3-0-i2i-250628"] = "Custom prompt for seededit-3-0-i2i-250628"

MAX_REFERENCE_IMAGES = 3
CUSTOM_CATEGORY_OPTION = "Others"
categories_list = [
    "Accessories & Supplies",
    "Additive Manufacturing Products",
    "Abrasive & Finishing Products",
    "Arts & Crafts Supplies",
    "Arts, Crafts & Sewing Storage",
    "Automotive Enthusiast Merchandise",
    "Automotive Exterior Accessories",
    "Automotive Interior Accessories",
    "Automotive Paint & Paint Supplies",
    "Automotive Performance Parts & Accessories",
    "Automotive Replacement Parts",
    "Automotive Tires & Wheels",
    "Automotive Tools & Equipment",
    "Baby",
    "Baby Activity & Entertainment Products",
    "Baby & Child Care Products",
    "Baby & Toddler Feeding Supplies",
    "Baby & Toddler Toys",
    "Baby Care Products",
    "Baby Diapering Products",
    "Baby Gifts",
    "Baby Girls' Clothing & Shoes",
    "Baby Boys' Clothing & Shoes",
    "Baby Safety Products",
    "Baby Stationery",
    "Baby Strollers & Accessories",
    "Baby Travel Gear",
    "Backpacks",
    "Bath Products",
    "Beading & Jewelry Making",
    "Bedding",
    "Beauty & Personal Care",
    "Beauty Tools & Accessories",
    "Boys' Accessories",
    "Boys' Clothing",
    "Boys' Jewelry",
    "Boys' School Uniforms",
    "Boys' Shoes",
    "Boys' Watches",
    "Building Supplies",
    "Building Toys",
    "Camera & Photo",
    "Car Care",
    "Car Electronics & Accessories",
    "Cat Supplies",
    "Cell Phones & Accessories",
    "Child Safety Car Seats & Accessories",
    "Commercial Door Products",
    "Computer Components",
    "Computer External Components",
    "Computer Monitors",
    "Computer Networking",
    "Computer Servers",
    "Computers",
    "Computers & Tablets",
    "Craft & Hobby Fabric",
    "Craft Supplies & Materials",
    "Cutting Tools",
    "Data Storage",
    "Diet & Sports Nutrition",
    "Dolls & Accessories",
    "Dog Supplies",
    "eBook Readers & Accessories",
    "Electrical Equipment",
    "Electronic Components",
    "Fabric Decorating",
    "Fasteners",
    "Filtration",
    "Fish & Aquatic Pets",
    "Finger Toys",
    "Food Service Equipment & Supplies",
    "Foot, Hand & Nail Care Products",
    "Furniture",
    "Games & Accessories",
    "Garment Bags",
    "Gift Cards",
    "Gift Wrapping Supplies",
    "Girls' Accessories",
    "Girls' Clothing",
    "Girls' Jewelry",
    "Girls' School Uniforms",
    "Girls' Shoes",
    "Girls' Watches",
    "GPS & Navigation",
    "Hair Care Products",
    "Hardware",
    "Health & Household",
    "Health Care Products",
    "Headphones & Earbuds",
    "Heavy Duty & Commercial Vehicle Equipment",
    "Heating, Cooling & Air Quality",
    "Home Appliances",
    "Home Audio & Theater Products",
    "Home Dcor Products",
    "Home Lighting & Ceiling Fans",
    "Home Storage & Organization",
    "Horse Supplies",
    "Household Cleaning Supplies",
    "Household Supplies",
    "Hydraulics, Pneumatics & Plumbing",
    "Industrial Adhesives, Sealants & Lubricants",
    "Industrial Hardware",
    "Industrial Materials",
    "Industrial Power & Hand Tools",
    "Industrial & Scientific",
    "Ironing Products",
    "Janitorial & Sanitation Supplies",
    "Kids' Dress Up & Pretend Play",
    "Kids' Electronics",
    "Kids' Furniture",
    "Kids' Home Store",
    "Kids' Party Supplies",
    "Kids' Play Boats",
    "Kids' Play Buses",
    "Kids' Play Cars & Race Cars",
    "Kids' Play Tractors",
    "Kids' Play Trains & Trams",
    "Kids' Play Trucks",
    "Kitchen & Bath Fixtures",
    "Kitchen & Dining",
    "Knitting & Crochet Supplies",
    "Lab & Scientific Products",
    "Laptop Accessories",
    "Laptop Bags",
    "Learning & Education Toys",
    "Legacy Systems",
    "Light Bulbs",
    "Lights, Bulbs & Indicators",
    "Luggage",
    "Luggage Sets",
    "Makeup",
    "Material Handling Products",
    "Measuring & Layout",
    "Men's Accessories",
    "Men's Clothing",
    "Men's Shoes",
    "Men's Watches",
    "Messenger Bags",
    "Motorcycle & Powersports",
    "Needlework Supplies",
    "Nintendo 3DS & 2DS Consoles, Games & Accessories",
    "Nintendo DS Games, Consoles & Accessories",
    "Nintendo Switch Consoles, Games & Accessories",
    "Novelty Toys & Amusements",
    "Nursery Furniture, Bedding & Dcor",
    "Occupational Health & Safety Products",
    "Office Electronics",
    "Oil & Fluids",
    "Online Video Game Services",
    "Oral Care Products",
    "Outdoor Recreation",
    "Packaging & Shipping Supplies",
    "Paint, Wall Treatments & Supplies",
    "Painting, Drawing & Art Supplies",
    "Party Decorations",
    "Party Supplies",
    "Perfumes & Fragrances",
    "Personal Care Products",
    "Pet Bird Supplies",
    "PlayStation 3 Games, Consoles & Accessories",
    "PlayStation 4 Games, Consoles & Accessories",
    "PlayStation 5 Consoles, Games & Accessories",
    "PlayStation Vita Games, Consoles & Accessories",
    "Portable Audio & Video",
    "Power Tools & Hand Tools",
    "Power Transmission Products",
    "Pregnancy & Maternity Products",
    "Printmaking Supplies",
    "Professional Dental Supplies",
    "Professional Medical Supplies",
    "Pumps & Plumbing Equipment",
    "Puppets & Puppet Theaters",
    "Puzzles",
    "Rain Umbrellas",
    "Reptiles & Amphibian Supplies",
    "Retail Store Fixtures & Equipment",
    "RV Parts & Accessories",
    "Safety & Security",
    "Science Education Supplies",
    "Scrapbooking & Stamping Supplies",
    "Security & Surveillance Equipment",
    "Seasonal Dcor",
    "Sewing Products",
    "Sexual Wellness Products",
    "Shaving & Hair Removal Products",
    "Skin Care Products",
    "Slot Cars, Race Tracks & Accessories",
    "Small Animal Supplies",
    "Smart Home - Heating & Cooling",
    "Smart Home Thermostats - Compatibility Checker",
    "Smart Home: Home Entertainment",
    "Smart Home: Lawn and Garden",
    "Smart Home: Lighting",
    "Smart Home: New Smart Devices",
    "Smart Home: Other Solutions",
    "Smart Home: Plugs and Outlets",
    "Smart Home: Security Cameras and Systems",
    "Smart Home: Smart Locks and Entry",
    "Smart Home: Vacuums and Mops",
    "Smart Home: Voice Assistants and Hubs",
    "Smart Home: WiFi and Networking",
    "Software & PC Games",
    "Sony PSP Games, Consoles & Accessories",
    "Sports & Fitness",
    "Sports & Outdoor Play Toys",
    "Sports & Outdoors",
    "Sports Nutrition Products",
    "Stationery & Gift Wrapping Supplies",
    "Stuffed Animals & Plush Toys",
    "Suitcases",
    "Tablet Accessories",
    "Tablet Replacement Parts",
    "Televisions & Video Products",
    "Test, Measure & Inspect",
    "Tools & Home Improvement",
    "Toilet Training Products",
    "Toy Figures & Playsets",
    "Toy Vehicle Playsets",
    "Tricycles, Scooters & Wagons",
    "Travel Accessories",
    "Travel Duffel Bags",
    "Travel Tote Bags",
    "Vehicle Electronics",
    "Video Game Consoles & Accessories",
    "Video Games",
    "Video Projectors",
    "Vision Products",
    "Virtual Reality Hardware & Accessories",
    "Wall Art",
    "Wearable Technology",
    "Wellness & Relaxation Products",
    "Welding & Soldering",
    "Wii Games, Consoles & Accessories",
    "Wii U Games, Consoles & Accessories",
    "Women's Accessories",
    "Women's Clothing",
    "Women's Handbags",
    "Women's Jewelry",
    "Women's Shoes",
    "Women's Watches",
    "Xbox 360 Games, Consoles & Accessories",
    "Xbox One Games, Consoles & Accessories",
    "Xbox Series X & S Consoles, Games & Accessories",
    CUSTOM_CATEGORY_OPTION,
]
CUSTOM_CATEGORY_MATCHES = {CUSTOM_CATEGORY_OPTION, "Other"}


def log_entry(entry: List[Any]) -> None:
    row = list(entry)
    while len(row) < len(EXCEL_HEADERS):
        row.append(None)
    row = row[: len(EXCEL_HEADERS)]
    payload = dict(zip(EXCEL_HEADERS, row))
    ts = str(payload.get("Timestamp") or datetime.now(ZoneInfo("UTC")).isoformat())
    safe_ts = ts.replace(" ", "_").replace(":", "-")
    key = f"{LOG_RUNS_PREFIX.rstrip('/')}/{safe_ts}_{uuid.uuid4().hex}.json"
    s3_client.put_object(
        Bucket=S3_BUCKET_NAME,
        Key=key,
        Body=json.dumps(payload, default=str).encode("utf-8"),
        ContentType="application/json",
    )
    append_row_to_excel(row)


def upload_image_to_s3(image_path: Path, key_prefix: str) -> Tuple[Optional[str], Optional[str]]:
    link = upload_file_to_s3(image_path, key_prefix, content_type="image/jpeg", public=True)
    return link, link


def decode_possible_base64(value: Any) -> Optional[bytes]:
    if value is None:
        return None
    if isinstance(value, (bytes, bytearray)):
        return bytes(value)
    if isinstance(value, str):
        for strict in (True, False):
            try:
                return base64.b64decode(value, validate=strict)
            except Exception:
                continue
    return None


def image_from_base64(data: Any) -> Optional[Image.Image]:
    raw = decode_possible_base64(data)
    if not raw:
        return None
    try:
        return Image.open(io.BytesIO(raw)).convert("RGB")
    except Exception as exc:
        print("[error] Unable to decode base64 image:", exc)
        return None


def pil_to_base64(image: Image.Image) -> str:
    buf = io.BytesIO()
    image.convert("RGB").save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def pil_to_data_uri(image: Image.Image, fmt: str = "png") -> str:
    buf = io.BytesIO()
    image.convert("RGB").save(buf, format=fmt.upper())
    encoded = base64.b64encode(buf.getvalue()).decode("utf-8")
    return f"data:image/{fmt.lower()};base64,{encoded}"


def _image_from_ark_response(resp_obj: Any) -> Optional[Image.Image]:
    if resp_obj is None:
        return None
    url = getattr(resp_obj, "url", None)
    if url:
        img = download_image_from_url(url)
        if img:
            return img
    b64_data = getattr(resp_obj, "b64_json", None)
    if b64_data:
        return image_from_base64(b64_data)
    return None


async def seedream_generate(
    prompt: str,
    image: Optional[Image.Image],
    image_urls: Optional[List[str]] = None,
    prefer_inline: bool = False,
) -> Optional[Image.Image]:
    if not ARK_API_KEY:
        print("[info] Seedream unavailable in this environment.")
        return None

    reference_urls = [url for url in (image_urls or []) if url and url.startswith(("http://", "https://"))]
    inline_reference = None
    if image is not None and (prefer_inline or not reference_urls):
        inline_reference = pil_to_data_uri(image)

    def _prepare_image_arg():
        if inline_reference:
            return inline_reference
        if not reference_urls:
            return None
        return reference_urls if len(reference_urls) > 1 else reference_urls[0]

    if ARK_SDK_CLIENT:
        def _sdk_call():
            request_kwargs = {
                "model": "seedream-4-0-250828",
                "prompt": prompt,
                "size": "2K",
                "response_format": "url",
                "watermark": False,
                "sequential_image_generation": "auto",
            }
            image_arg = _prepare_image_arg()
            if image_arg is not None:
                request_kwargs["image"] = image_arg
            if SequentialImageGenerationOptions:
                request_kwargs["sequential_image_generation_options"] = SequentialImageGenerationOptions(max_images=1)
            resp = ARK_SDK_CLIENT.images.generate(**request_kwargs)
            return _image_from_ark_response(resp.data[0] if resp.data else None)

        return await run_blocking(_sdk_call)

    def _http_call():
        payload = {
            "model": "seedream-4-0-250828",
            "prompt": prompt,
            "size": "2K",
            "response_format": "url",
            "watermark": True,
            "sequential_image_generation": "auto",
        }
        image_arg = _prepare_image_arg()
        if image_arg is not None:
            payload["image"] = image_arg
        headers = {
            "Authorization": f"Bearer {ARK_API_KEY}",
            "Content-Type": "application/json",
        }
        try:
            resp = requests.post(SEEDREAM_ENDPOINT, json=payload, headers=headers, timeout=120)
        except Exception as exc:
            print("[error] Seedream API request failed:", exc)
            return None
        if resp.status_code != 200:
            print(f"[error] Seedream HTTP {resp.status_code}: {resp.text}")
            return None
        try:
            data = resp.json()
        except ValueError as exc:
            print("[error] Seedream response was not JSON:", exc, resp.text)
            return None
        images = data.get("data")
        if not images:
            print("[error] Seedream response missing image data:", data)
            return None
        first = images[0]
        url = first.get("url")
        if url:
            return download_image_from_url(url)
        b64_data = first.get("b64_json")
        if b64_data:
            return image_from_base64(b64_data)
        print("[error] Seedream response missing usable image data:", first)
        return None

    return await run_blocking(_http_call)



async def seededit_generate(
    prompt: str,
    image: Optional[Image.Image],
    reference_url: Optional[str],
    prefer_inline: bool = False,
) -> Optional[Image.Image]:
    if not ARK_API_KEY:
        print("[info] SeedEdit unavailable in this environment.")
        return None
    if not ARK_SDK_CLIENT:
        print("[info] SeedEdit requires the Ark SDK client.")
        return None
    if reference_url is None and image is None:
        print("[warn] SeedEdit requires at least one reference image.")
        return None

    inline_reference = None
    if image is not None and (prefer_inline or not reference_url):
        inline_reference = pil_to_data_uri(image)

    def _call():
        request_kwargs = {
            "model": "seededit-3-0-i2i-250628",
            "prompt": prompt,
            "response_format": "url",
            "size": "adaptive",
            "seed": 123,
            "guidance_scale": 5.5,
            "watermark": True,
        }
        if inline_reference:
            request_kwargs["image"] = inline_reference
        elif reference_url:
            request_kwargs["image"] = reference_url
        elif image is not None:
            request_kwargs["image"] = pil_to_data_uri(image)
        resp = ARK_SDK_CLIENT.images.generate(**request_kwargs)
        return _image_from_ark_response(resp.data[0] if resp.data else None)

    return await run_blocking(_call)


def build_reference_url(image: Optional[Image.Image], link: Optional[str]) -> Optional[str]:
    if link:
        return link
    if image is not None:
        return pil_to_data_uri(image)
    return None


def add_reference_to_prompt(prompt: str, reference_url: Optional[str]) -> str:
    if not reference_url:
        return prompt
    return (
        f"{prompt}\n\nReference Image URL: {reference_url}\n"
        "Preserve the original product exactly while embedding it naturally in the new scene."
    )


def to_pil_image(image: Any) -> Optional[Image.Image]:
    if image is None:
        return None
    if isinstance(image, Image.Image):
        return image
    if isinstance(image, np.ndarray):
        if image.dtype != "uint8":
            image = image.astype("uint8")
        return Image.fromarray(image)
    return None


def combine_reference_images(images: List[Image.Image]) -> Image.Image:
    """
    Merge multiple reference shots side-by-side so models receive a single contextual image.
    """
    if not images:
        raise ValueError("No images provided for combination.")
    if len(images) == 1:
        return images[0]

    max_height = max(img.height for img in images)
    normalized: List[Image.Image] = []
    for img in images:
        if img.height != max_height:
            scale = max_height / img.height
            new_width = max(1, int(img.width * scale))
            img = img.resize((new_width, max_height), Image.LANCZOS)
        normalized.append(img)

    total_width = sum(img.width for img in normalized)
    combined = Image.new("RGB", (total_width, max_height), color=(255, 255, 255))
    offset = 0
    for img in normalized:
        combined.paste(img, (offset, 0))
        offset += img.width
    return combined


def ensure_aspect_ratio_bounds(
    image: Optional[Image.Image],
    min_ratio: float = 0.34,
    max_ratio: float = 2.9,
) -> Tuple[Optional[Image.Image], bool]:
    if image is None:
        return None, False
    width, height = image.size
    if not width or not height:
        return image, False
    ratio = width / height
    if min_ratio <= ratio <= max_ratio:
        return image, False
    target_width = width
    target_height = height
    if ratio < min_ratio:
        target_width = max(width, math.ceil(height * min_ratio))
    else:
        target_height = max(height, math.ceil(width / max_ratio))
    canvas = Image.new("RGB", (target_width, target_height), color=(255, 255, 255))
    offset_x = (target_width - width) // 2
    offset_y = (target_height - height) // 2
    canvas.paste(image.convert("RGB"), (offset_x, offset_y))
    return canvas, True

def save_input_image(image: Optional[Image.Image], ts: str) -> Tuple[Optional[str], Optional[str]]:
    if image is None:
        return None, None
    temp_path = Path(tempfile.gettempdir()) / f"{ts}_input.jpg"
    image.convert("RGB").save(temp_path, format="JPEG", quality=95)
    download_link, view_link = upload_image_to_s3(temp_path, S3_INPUT_PREFIX)
    temp_path.unlink(missing_ok=True)
    return download_link, view_link


def download_image_from_url(url: str) -> Optional[Image.Image]:
    if not url:
        return None
    try:
        response = requests.get(url, timeout=15)
        response.raise_for_status()
        return Image.open(io.BytesIO(response.content)).convert("RGB")
    except Exception as exc:
        print("[warn] Failed to download image:", exc)
        return None




def _extract_first_image_from_output(output: Any) -> Optional[Image.Image]:
    """Return the first usable image from Replicate outputs (urls, base64, nested)."""
    if output is None:
        return None
    if isinstance(output, dict):
        url = output.get("url")
        if url:
            img = download_image_from_url(url)
            if img:
                return img
        data = output.get("b64_json") or output.get("base64")
        if data:
            img = image_from_base64(data)
            if img:
                return img
        return _extract_first_image_from_output(output.get("output"))
    if isinstance(output, (list, tuple)):
        for item in output:
            img = _extract_first_image_from_output(item)
            if img:
                return img
        return None
    if isinstance(output, str):
        return download_image_from_url(output) or image_from_base64(output)
    return None


async def generate_with_openai_image_model(
    model_name: str,
    prompt: str,
    image: Optional[Image.Image],
    input_link: Optional[str],
) -> Optional[Image.Image]:
    reference_url = build_reference_url(image, input_link)
    prompt_with_reference = add_reference_to_prompt(prompt, reference_url)
    prompt_with_reference = clamp_prompt_length(prompt_with_reference)

    def _call():
        response = OPENAI_CLIENT.images.generate(
            model=model_name,
            prompt=prompt_with_reference,
            size="1024x1024",
        )
        data = response.data[0].b64_json
        return image_from_base64(data)

    return await run_blocking(_call)


async def generate_with_gemini_image_model(
    model_name: str,
    prompt: str,
    image: Optional[Image.Image],
    input_link: Optional[str],
) -> Optional[Image.Image]:
    reference_url = build_reference_url(image, input_link)
    prompt = clamp_prompt_length(prompt)

    def _call():
        contents: List[Any] = [prompt]
        if image is not None:
            contents.append(image)
        elif reference_url:
            downloaded = download_image_from_url(reference_url)
            if downloaded:
                contents.append(downloaded)
        resp = gclient.models.generate_content(
            model=model_name,
            contents=contents,
        )
        candidates = getattr(resp, "candidates", None) or []
        for candidate in candidates:
            candidate_content = getattr(candidate, "content", None)
            if not candidate_content:
                continue
            for part in getattr(candidate_content, "parts", None) or []:
                inline_data = getattr(part, "inline_data", None)
                if inline_data and getattr(inline_data, "data", None):
                    return image_from_base64(inline_data.data)
        resp_parts = getattr(resp, "parts", None)
        if resp_parts:
            for part in resp_parts:
                inline_data = getattr(part, "inline_data", None)
                if inline_data and getattr(inline_data, "data", None):
                    return image_from_base64(inline_data.data)
        return None

    return await run_blocking(_call)


async def generate_with_seedream_model(
    model_name: str,
    prompt: str,
    image: Optional[Image.Image],
    input_link: Optional[str],
) -> Optional[Image.Image]:
    prepared_image, was_adjusted = ensure_aspect_ratio_bounds(image)
    image_urls = [input_link] if input_link else None
    prefer_inline = was_adjusted or not image_urls
    return await seedream_generate(
        prompt,
        prepared_image,
        image_urls=image_urls,
        prefer_inline=prefer_inline,
    )


async def generate_with_seededit_model(
    model_name: str,
    prompt: str,
    image: Optional[Image.Image],
    input_link: Optional[str],
) -> Optional[Image.Image]:
    prepared_image, was_adjusted = ensure_aspect_ratio_bounds(image)
    reference_url = build_reference_url(prepared_image, input_link)
    prefer_inline = was_adjusted or not reference_url
    return await seededit_generate(
        prompt,
        prepared_image,
        reference_url,
        prefer_inline=prefer_inline,
    )


async def generate_with_replicate_model(
    model_name: str,
    prompt: str,
    image: Optional[Image.Image],
    input_link: Optional[str],
) -> Optional[Image.Image]:
    token = REPLICATE_API_TOKEN
    if not token:
        print(f"[info] Replicate token missing; skipping {model_name}.")
        return None
    prepared_image, _ = ensure_aspect_ratio_bounds(image)
    reference_url = build_reference_url(prepared_image, input_link)
    image_arg = None
    if prepared_image is not None:
        image_arg = pil_to_data_uri(prepared_image)
    elif reference_url:
        image_arg = reference_url

    model_id = REPLICATE_MODEL_IDS.get(model_name, model_name)

    def _build_input() -> Optional[Dict[str, Any]]:
        if model_name == "qwen/qwen-image-edit-plus":
            imgs = []
            if image_arg:
                imgs = [image_arg, image_arg]
            elif reference_url:
                imgs = [reference_url, reference_url]
            if not imgs:
                print(f"[warn] {model_name} requires two images; none provided.")
                return None
            return {"image": imgs, "prompt": prompt}
        if model_name == "qwen/qwen-image-edit":
            img = image_arg or reference_url
            if not img:
                print(f"[warn] {model_name} requires an image; none provided.")
                return None
            return {"image": img, "prompt": prompt, "output_quality": 80}
        if model_name in {"black-forest-labs/flux-kontext-max", "black-forest-labs/flux-kontext-pro"}:
            img = image_arg or reference_url
            if not img:
                print(f"[warn] {model_name} requires an image; none provided.")
                return None
            return {"prompt": prompt, "input_image": img, "output_format": "jpg"}
        if model_name == "black-forest-labs/flux-1.1-pro":
            return {"prompt": prompt, "prompt_upsampling": True}
        if model_name == "black-forest-labs/flux-1.1-pro-ultra":
            return {"prompt": prompt, "aspect_ratio": "3:2"}
        if model_name == "black-forest-labs/flux-pro":
            return {"prompt": prompt}
        if model_name == "ideogram-ai/ideogram-v3-turbo":
            return {"prompt": prompt, "aspect_ratio": "3:2"}
        if model_name == "qwen/qwen-image":
            return {"prompt": prompt, "guidance": 4, "num_inference_steps": 30}
        payload = {"prompt": prompt}
        if image_arg:
            payload["image"] = image_arg
        return payload

    payload_input = _build_input()
    if payload_input is None:
        return None

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    create_url = f"https://api.replicate.com/v1/models/{model_id}/predictions"

    def _call():
        try:
            resp = requests.post(create_url, json={"input": payload_input}, headers=headers, timeout=30)
        except Exception as exc:  # noqa: BLE001
            print(f"[warn] Replicate request failed for {model_name}:", exc)
            return None
        if resp.status_code not in (200, 201, 202):
            print(f"[warn] Replicate {model_name} HTTP {resp.status_code}: {resp.text}")
            return None
        data = resp.json()
        status = data.get("status")
        image_out = _extract_first_image_from_output(data.get("output"))
        poll_url = data.get("urls", {}).get("get")
        pred_id = data.get("id")
        if not poll_url and pred_id:
            poll_url = f"https://api.replicate.com/v1/predictions/{pred_id}"
        terminal = {"succeeded", "failed", "canceled"}
        while image_out is None and status not in terminal and poll_url:
            try:
                poll_resp = requests.get(poll_url, headers=headers, timeout=30)
            except Exception as exc:  # noqa: BLE001
                print(f"[warn] Replicate poll failed for {model_name}:", exc)
                break
            if poll_resp.status_code not in (200, 201, 202):
                print(f"[warn] Replicate poll HTTP {poll_resp.status_code}: {poll_resp.text}")
                break
            payload = poll_resp.json()
            status = payload.get("status", status)
            image_out = _extract_first_image_from_output(payload.get("output"))
            if image_out or status in terminal:
                break
            time.sleep(2)
        if image_out is None:
            print(f"[warn] Replicate {model_name} returned no image (status={status}).")
        return image_out

    return await run_blocking(_call)


MODEL_GENERATORS: Dict[str, Any] = {
    "gpt-image-1": generate_with_openai_image_model,
    "gpt-image-1-mini": generate_with_openai_image_model,
    "gemini-2.5-flash-image": generate_with_gemini_image_model,
    "gemini-3-pro-image-preview": generate_with_gemini_image_model,
    "qwen/qwen-image-edit": generate_with_replicate_model,
    "qwen/qwen-image-edit-plus": generate_with_replicate_model,
    "black-forest-labs/flux-kontext-max": generate_with_replicate_model,
    "black-forest-labs/flux-kontext-pro": generate_with_replicate_model,
    "black-forest-labs/flux-pro": generate_with_replicate_model,
    "ideogram-ai/ideogram-v3-turbo": generate_with_replicate_model,
    "qwen/qwen-image": generate_with_replicate_model,
    "seedream-4": generate_with_seedream_model,
    "seededit-3-0-i2i-250628": generate_with_seededit_model,
}


async def generate_for_model(
    model_name: str,
    category: str,
    description: str,
    image: Optional[Image.Image],
    ts: str,
    input_download_link: Optional[str],
    input_view_link: Optional[str],
    output_index: int = 1,
    model_prompts: Optional[Dict[str, str]] = None,
) -> Optional[Image.Image]:
    template = PROMPT_TEMPLATE_SEEDREAM if model_name == "seedream-4" else PROMPT_TEMPLATE
    base_prompt = template.format(category=category, description=description)
    custom_prompt = (model_prompts or {}).get(model_name)
    prompt = clamp_prompt_length(custom_prompt if custom_prompt else base_prompt)
    generator = MODEL_GENERATORS.get(model_name)
    if not generator:
        raise ValueError(f"Unsupported model: {model_name}")
    inference_start = time.perf_counter()
    out_img = await generator(model_name, prompt, image, input_download_link)
    if not out_img:
        print(f"[warn] {model_name} returned no image.")
        return None

    safe_model_name = model_name.replace("/", "_").replace(":", "-")
    temp_path = Path(tempfile.gettempdir()) / f"{ts}_{safe_model_name}_gen{output_index}.jpg"
    out_img.save(temp_path)
    out_download_link, out_view_link = upload_image_to_s3(temp_path, S3_OUTPUT_PREFIX)
    temp_path.unlink(missing_ok=True)
    latency = round(time.perf_counter() - inference_start, 3)
    speed = round(1 / latency, 3) if latency else None
    tokens = len(prompt.split())
    resolution = f"{out_img.width}x{out_img.height}"

    log_entry([
        ts,
        category,
        description,
        model_name,
        output_index,
        None,
        None,
        None,
        input_view_link,
        input_download_link,
        out_view_link,
        out_download_link,
        speed,
        latency,
        tokens,
        1,
        resolution,
    ])

    return out_img


async def generate_all(
    category: str,
    description: str,
    image: Optional[Image.Image],
    output_index: int = 1,
    model_prompts: Optional[Dict[str, str]] = None,
) -> List[Optional[Image.Image]]:
    ts = datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%Y-%m-%d_%H-%M-%S")
    pil_image = to_pil_image(image)
    if pil_image is None:
        return [None] * len(MODELS)
    input_download_link, input_view_link = save_input_image(pil_image, ts)
    tasks = [
        generate_for_model(
            model,
            category,
            description,
            pil_image,
            ts,
            input_download_link,
            input_view_link,
            output_index,
            model_prompts=model_prompts,
        )
        for model in MODELS
    ]
    results = await asyncio.gather(*tasks, return_exceptions=True)
    formatted: List[Optional[Image.Image]] = []
    for result in results:
        if isinstance(result, Exception):
            print("[error] Generation failed:", result)
            formatted.append(None)
        else:
            formatted.append(result)
    return formatted


def generate_sync(
    category: str,
    description: str,
    image: Any,
    output_index: int = 1,
    model_prompts: Optional[Dict[str, str]] = None,
) -> List[Optional[Image.Image]]:
    with GENERATION_LOCK:
        try:
            return asyncio.run(generate_all(category, description, image, output_index, model_prompts=model_prompts))
        except RuntimeError as exc:
            message = str(exc).lower()
            if "asyncio.run()" in message or "event loop is already running" in message:
                def _runner() -> List[Optional[Image.Image]]:
                    return asyncio.run(generate_all(category, description, image, output_index, model_prompts=model_prompts))

                with ThreadPoolExecutor(max_workers=1) as executor:
                    future = executor.submit(_runner)
                    return future.result()
            raise


def handle_feedback(model_id: str, image_index: int, feedback: str, text: str, score: Optional[float]) -> str:
    ts = datetime.now(ZoneInfo("UTC")).isoformat()
    safe_ts = ts.replace(" ", "_").replace(":", "-")
    payload = {
        "timestamp": ts,
        "model_id": model_id,
        "image_index": image_index,
        "feedback": feedback,
        "review": text,
        "score": score,
    }
    key = f"{LOG_FEEDBACK_PREFIX.rstrip('/')}/{safe_ts}_{model_id}_{uuid.uuid4().hex}.json"
    s3_client.put_object(
        Bucket=S3_BUCKET_NAME,
        Key=key,
        Body=json.dumps(payload, default=str).encode("utf-8"),
        ContentType="application/json",
    )

    with EXCEL_LOCK:
        tmp = download_excel_from_s3()
        try:
            df = pd.read_excel(tmp)
            expected_cols = [
                "Timestamp",
                "Category",
                "Description",
                "Model Name",
                "Output Image No",
                "Thumbs (1=Like,0=Dislike)",
                "Review",
                "Score by Human",
                "Input Image View Link",
                "Input Image Download Link",
                "Output Image View Link",
                "Output Image Download Link",
                "Inference Speed (img/sec)",
                "Latency (sec)",
                "Token Processed (in+out)",
                "Images Generated",
                "Image Quality (resolution)",
            ]
            df = df.reindex(columns=expected_cols)
            match = df[(df["Model Name"] == model_id) & (df["Output Image No"] == image_index)]
            if match.empty:
                return " No matching entry found."
            idx = match.index[-1]
            df.loc[idx, "Thumbs (1=Like,0=Dislike)"] = 1 if feedback == "up" else 0
            df.loc[idx, "Review"] = text
            df.loc[idx, "Score by Human"] = score
            df.to_excel(tmp, index=False)
            upload_excel_to_s3(tmp)
        finally:
            tmp.unlink(missing_ok=True)

    return f" Feedback saved for {model_id} (image {image_index})"


def load_uploaded_image(uploaded_file: Optional[Any]) -> Optional[Image.Image]:
    if uploaded_file is None:
        return None
    try:
        return Image.open(uploaded_file).convert("RGB")
    except Exception as exc:  # noqa: BLE001
        print("[warn] Failed to read uploaded image:", exc)
        return None


def update_status_text(value: str) -> None:
    st.session_state["status_text"] = value


def reset_session_state() -> None:
    """Clear all Streamlit session state and rerun to start fresh."""
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    try:
        st.experimental_rerun()  # Streamlit <=1.31
    except AttributeError:
        st.rerun()  # Streamlit >=1.32


def main() -> None:
    st.set_page_config(page_title="Infographics Product Image Generator", layout="wide")
    st.markdown("#  Infographics Product Image Generator & Feedback (S3 Storage)")

    if "generation" not in st.session_state:
        st.session_state["generation"] = None
    if "status_text" not in st.session_state:
        st.session_state["status_text"] = "Ready!"

    if st.button("Reset"):
        reset_session_state()

    with st.form("generator_form"):
        st.write("### Reference Images")
        upload_cols = st.columns(MAX_REFERENCE_IMAGES)
        uploads: List[Any] = []
        url_inputs: List[str] = []
        for idx, col in enumerate(upload_cols, start=1):
            with col:
                uploads.append(col.file_uploader(f"Upload Image {idx}", type=["png", "jpg", "jpeg"], key=f"upload_{idx}"))
                url_inputs.append(col.text_input(f"Or Paste Image {idx} URL", key=f"url_{idx}"))

        category = st.selectbox("Select Category", options=categories_list, index=0, key="category_select")
        custom_cat_value = st.text_input(
            f"Custom Category (used when {CUSTOM_CATEGORY_OPTION} selected)",
            key="custom_category",
            placeholder="Type your custom category here",
        )
        use_custom_category = category in CUSTOM_CATEGORY_MATCHES
        description = st.text_area("Description", key="description_text", height=80)
        submitted = st.form_submit_button(" Generate Infographics")

    if submitted:
        collected: List[Image.Image] = []
        for upload, url in zip(uploads, url_inputs):
            img = load_uploaded_image(upload)
            if img is None and url:
                img = download_image_from_url(url.strip())
            if img is not None:
                collected.append(img)

        model_prompts = {
            name: val.strip()
            for name, val in MODEL_PROMPT_OVERRIDES.items()
            if val and val.strip()
        }

        final_category = (
            custom_cat_value.strip()
            if use_custom_category and custom_cat_value
            else category
        ) or CUSTOM_CATEGORY_OPTION
        if not collected:
            update_status_text(" Please upload at least one image or provide a valid image URL.")
            st.warning("Please upload at least one image or provide a valid image URL.")
        else:
            try:
                reference_image = combine_reference_images(collected)
            except Exception as exc:  # noqa: BLE001
                update_status_text(" Unable to prepare the reference preview from the provided inputs.")
                st.error(f"Unable to prepare the reference preview: {exc}")
            else:
                generated_images = generate_sync(
                    final_category,
                    description,
                    reference_image,
                    output_index=1,
                    model_prompts=model_prompts or None,
                )
                produced_outputs = sum(1 for img in generated_images if img is not None)
                status_msg = (
                    f" Combined {len(collected)} reference image(s) & generated {produced_outputs} outputs."
                    if produced_outputs
                    else " Generation failed for the provided references."
                )
                update_status_text(status_msg)
                st.session_state["generation"] = {
                    "reference_image": reference_image,
                    "outputs": generated_images,
                    "category": final_category,
                    "description": description,
                }

    st.text_input("Status", value=st.session_state["status_text"], disabled=True)

    generation = st.session_state.get("generation")
    if generation:
        st.markdown("### Reference Preview")
        ref_cols = st.columns([1, 2, 1])
        with ref_cols[1]:
            st.image(generation["reference_image"], caption="Combined Reference", width=380)

        st.markdown("### Model Outputs")
        for model_name, img in zip(MODELS, generation["outputs"]):
            st.markdown(f"#### {model_name}")
            if img is None:
                st.info("No image generated for this model.")
                continue

            col_image, col_react, col_review, col_score, col_save = st.columns([3, 1.2, 3, 1.2, 1])
            review_key = f"review_{model_name}"
            score_key = f"score_{model_name}"
            feedback_key = f"feedback_status_{model_name}"
            choice_key = f"feedback_choice_{model_name}"
            if choice_key not in st.session_state:
                st.session_state[choice_key] = None

            with col_image:
                st.image(img, use_container_width=True)

            with col_react:
                st.write("Reaction")
                up_clicked = st.button("👍", key=f"thumb_up_{model_name}")
                down_clicked = st.button("👎", key=f"thumb_down_{model_name}")
                if up_clicked:
                    st.session_state[choice_key] = "up"
                if down_clicked:
                    st.session_state[choice_key] = "down"
                selection = st.session_state.get(choice_key)
                st.caption(f"Selected: {selection or 'None'}")

            with col_review:
                st.text_area("Review", key=review_key, height=110)

            with col_score:
                st.number_input(
                    "Score (0-100)",
                    min_value=0,
                    max_value=100,
                    step=1,
                    key=score_key,
                )

            with col_save:
                save_clicked = st.button("Save", key=f"save_{model_name}")
                if save_clicked:
                    selected = st.session_state.get(choice_key)
                    if not selected:
                        st.warning("Select 👍 or 👎 first.")
                    else:
                        review_val = st.session_state.get(review_key, "")
                        score_val = st.session_state.get(score_key)
                        message = handle_feedback(model_name, 1, selected, review_val, score_val)
                        st.session_state[feedback_key] = message

                if feedback_key in st.session_state:
                    st.success(st.session_state[feedback_key])


if __name__ == "__main__":
    main()
